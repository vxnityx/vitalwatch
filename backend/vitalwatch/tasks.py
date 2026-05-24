import csv
import os
from pathlib import Path

from celery import shared_task
from django.db import transaction
from openpyxl import load_workbook

from .models import StudentRecord, FacultyRecord
from .serializers import StudentRecordSerializer, FacultyRecordSerializer


STUDENT_FIELDS = [
    "College",
    "Course",
    "Year_Level",
    "Month",
    "Day",
    "Year",
    "StudentID",
    "Body_Temperature_C",
    "Blood_Pressure",
    "Heart_Rate_bpm",
    "Emotion",
]

FACULTY_FIELDS = [
    "College",
    "User_Type",
    "Month",
    "Day",
    "Year",
    "EmployeeID_or_Guest",
    "Body_Temperature_C",
    "Blood_Pressure",
    "Systolic_BP",
    "Diastolic_BP",
    "Heart_Rate_bpm",
    "Emotion",
    "Risk_Level",
    "Alert_Status",
]


def _field_map(expected_fields):
    return {field.lower(): field for field in expected_fields}


def _normalize_row(row, expected_fields):
    mapping = _field_map(expected_fields)
    normalized = {}

    for key, value in row.items():
        if key is None:
            continue
        field_name = mapping.get(key.strip().lower())
        if field_name:
            normalized[field_name] = value.strip() if isinstance(value, str) else value

    return normalized


def _sniff_dialect(sample_text):
    try:
        return csv.Sniffer().sniff(sample_text, delimiters=",\t;")
    except csv.Error:
        return csv.get_dialect("excel")


def _normalize_header(value):
    return value.strip().lower() if isinstance(value, str) else value


def _load_rows_from_csv(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as source_file:
        sample = source_file.read(4096)
        source_file.seek(0)
        dialect = _sniff_dialect(sample)
        reader = csv.DictReader(source_file, dialect=dialect)
        headers = reader.fieldnames or []
        rows = list(reader)
    return headers, rows


def _load_rows_from_xlsx(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = []
    for raw_row in rows[1:]:
        data_rows.append({headers[index]: value for index, value in enumerate(raw_row) if index < len(headers)})

    workbook.close()
    return headers, data_rows


def _load_rows_from_file(file_path):
    extension = Path(file_path).suffix.lower()

    if extension in {".csv", ".tsv", ".txt"}:
        return _load_rows_from_csv(file_path)

    if extension in {".xlsx", ".xlsm"}:
        return _load_rows_from_xlsx(file_path)

    raise ValueError("Unsupported file type. Use CSV, TSV, TXT, or XLSX.")


def _import_rows(model, serializer_class, expected_fields, file_path, task):
    created = 0
    processed = 0
    batch = []
    batch_size = 500

    with transaction.atomic():
        headers, rows = _load_rows_from_file(file_path)

        if not headers:
            raise ValueError("The uploaded file must include a header row.")

        missing_fields = [
            field
            for field in expected_fields
            if field.lower() not in {_normalize_header(name) for name in headers if name}
        ]
        if missing_fields:
            raise ValueError(f"Missing required columns: {', '.join(missing_fields)}")

        total_rows = len(rows)

        for row in rows:
            processed += 1
            normalized_row = _normalize_row(row, expected_fields)
            serializer = serializer_class(data=normalized_row)
            serializer.is_valid(raise_exception=True)
            batch.append(model(**serializer.validated_data))

            if len(batch) >= batch_size:
                model.objects.bulk_create(batch, batch_size=batch_size)
                created += len(batch)
                batch.clear()
                percentage = int((processed / total_rows) * 100) if total_rows else 100
                task.update_state(
                    state="PROGRESS",
                    meta={"processed": processed, "created": created, "total": total_rows, "percentage": percentage},
                )

        if batch:
            model.objects.bulk_create(batch, batch_size=batch_size)
            created += len(batch)

        percentage = 100 if total_rows else 100
        task.update_state(
            state="PROGRESS",
            meta={"processed": processed, "created": created, "total": total_rows, "percentage": percentage},
        )

    return {"processed": processed, "created": created, "record_type": model.__name__.lower()}


@shared_task(bind=True)
def import_records_from_file(self, file_path, record_type):
    try:
        if record_type == "student":
            result = _import_rows(StudentRecord, StudentRecordSerializer, STUDENT_FIELDS, file_path, self)
        elif record_type == "faculty":
            result = _import_rows(FacultyRecord, FacultyRecordSerializer, FACULTY_FIELDS, file_path, self)
        else:
            raise ValueError("Unsupported record type.")

        return {
            **result,
            "status": "completed",
        }
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@shared_task(name="app.tasks.process_trucking_upload", bind=True)
def process_trucking_upload(self, file_path, record_type=None, upload_id=None):
    inferred_type = record_type

    if inferred_type not in {"student", "faculty"}:
        headers, _ = _load_rows_from_file(file_path)
        normalized_headers = {_normalize_header(header) for header in headers if header}

        student_match_count = sum(1 for field in STUDENT_FIELDS if field.lower() in normalized_headers)
        faculty_match_count = sum(1 for field in FACULTY_FIELDS if field.lower() in normalized_headers)
        inferred_type = "faculty" if faculty_match_count > student_match_count else "student"

    return import_records_from_file.run(file_path, inferred_type)